import json
import openpyxl
from unicodedata import normalize
from user import User
from os import path
from datetime import datetime

# Estoque dos blocoquinhos do Kumon Passaré


class Bloco():
    """Classe Bloco representa os blocos do Kumon, e tem como atributos:
    user, representando a classe User, curso[Português, Matemática, inglês]; 
    estágio(varia de acordo com o curso);numero([001, 191], definidos de 10 em 10, 
    com exceção do estágio 7a ao 5a de Portugues); E por fim, quantidade, 
    que representa o número de blocos no estoque."""

    def __init__(self, user, curso, estagio, numero):
        self.user = user
        self.curso = curso
        self.estagio = estagio
        self.numero = numero
        self.quantidade = None
        # Obj None é substituido pelo valor refente a sua instancia, no setter.quantidade

    @staticmethod
    def getValoresJson(curso):
        # Função auxiliar do método relatorio, que obtem e retorna as infos dos arquivos json
        listaCurso = []
        try:
            with open(path.join(".", "models", f"{curso}.json"), "r+") as js:
                # Módulo path usado para garantir funcionamento em windows e linux
                dados = json.load(js)
                for estagio in dados.keys(): # Percorre json e obtém as quantidades de cada bloco
                    listaEstagio = []
                    for quantia in dados[estagio].values():
                        listaEstagio.append(quantia)
                    listaCurso.append(listaEstagio)
        except FileNotFoundError:
            print(f"Erro: Arquivo '{curso}.json' não encontrado!")
            return
        return listaCurso

    @classmethod
    # met é o atributo responsável por verificar se o método foi chamado pela classe ou instancia.
    def relatorio(cls, met=False):
        # Gera um relatório em formato XLSX(Excel) da quantidade de blocos registradas nos jsons
        wb = openpyxl.load_workbook("./models/Base relatorio.xlsx")
        primeirasLinhas = (5, 34, 65, 96)
        # Linhas iniciais arquivo XLSX
        slicesConjuntos = ((0, 7), (7, 14), (14, 21), (21, 30))
        # Pedaços de listaCurso em cada conjunto
        for sheet in wb.sheetnames:
            if sheet == "Portugues":
                conjuntosColunasXLSX = 4
            else:
                conjuntosColunasXLSX = 3
            estagio = cls.getValoresJson(sheet)
            planilha = wb[sheet]
            for conjuntos in range(conjuntosColunasXLSX):
                cont = 2
                primeiraLinha = primeirasLinhas[conjuntos]
                slice_estagio = estagio[slicesConjuntos[conjuntos][0]:
                                        slicesConjuntos[conjuntos][1]]
                for lista in slice_estagio:
                    for j in range(len(lista)):
                        planilha.cell(
                            row=j + primeiraLinha, column=cont, value=lista[j])
                    cont += 2
        wb.save("Relatorio Atualizado.xlsx")
        if met:
            return
        else:
            print("Relatório gerado com sucesso!")

    @property
    def curso(self):
        # Retorna o atributo curso
        return self._curso

    @curso.setter
    def curso(self, curso):
        # Valida e formata a atribuição do atributo curso
        curso = normalize('NFKD', curso).encode(
            'ASCII', 'ignore').decode('ASCII').lower().title()
        if curso not in ["Portugues", "Matematica", "Ingles"]:
            print("Curso Inválido!")
        else:
            self._curso = curso

    @property
    def estagio(self):
        # Retorna o atributo estagio
        return self._estagio

    @estagio.setter
    def estagio(self, estagio):
        # Valida e formata o atributo estagio
        estagio = estagio.upper()
        if self._curso == "Portugues":
            if estagio not in [
                    "7A", "6A", "5A", "4A", "3A", "2A", "A1", "A2", "B1", "B2",
                    "C1", "C2", "D1", "D2", "E1", "E2", "F1", "F2", "G1", "G2",
                    "H1", "H2", "I1", "I2", "J", "K", "L"]:
                print("Curso Inválido!")
            else:
                self._estagio = estagio
        elif self._curso == "Ingles":
            if estagio not in [
                    "7A", "6A", "5A", "4A", "3A", "2A", "A", "B", "C", "D",
                    "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]:
                print("Curso Inválido!")
            else:
                self._estagio = estagio
        else:
            if estagio not in [
                    "6A", "5A", "4A", "3A", "2A", "A", "B", "C", "D", "E", "F",
                    "G", "H", "I", "J", "K", "L", "M", "N", "O"]:
                print("Curso Inválido!")
            else:
                self._estagio = estagio

    @property
    def numero(self):
        # Retorna o atributo numero
        return self._numero

    @numero.setter
    def numero(self, numero):
        # Valida e formata o atributo numero
        numero = str(numero).title()
        if len(numero) < 3:
            while len(numero) != 3:
                numero = "0" + numero
        if self._curso == "Portugues" and self._estagio in ["7A", "6A", "5A"]:
            if numero not in ["Livro1", "Livro2", "Livro3", "Livro4"]:
                print("Número Inválido!")
            else:
                self._numero = numero
        else:
            if numero not in [
                    "001", "011", "021", "031", "041", "051", "061", "071",
                    "081", "091", "101", "111", "121", "131", "141", "151",
                    "161", "171", "181", "191"]:
                print("Número Inválido!")
            else:
                self._numero = numero

    @property
    def quantidade(self):
        # Retorna o atributo quantidade
        return self._quantidade

    @quantidade.setter
    def quantidade(self, quantidade):
        # Define o atributo quantidade a partir do arquivo json
        try:
            with open(path.join(".", "models", f"{self.curso}.json"), "r") as js:
                dados = json.load(js)
                self._quantidade = dados[self._estagio][self._numero]
        except FileNotFoundError:
            print(f"Erro: Arquivo '{self._curso}.json' não encontrado!")

    def adicionar(self, valor):
        # Adiciona ao atributo quantidade o valor passado como parâmetro e atualiza o json
        if self.user.log:
            # User precisa estar logado para realizar adição
            caminho = path.join(".", "models", f"{self.curso}.json")
            valor = abs(int(valor))
            try:
                with open(caminho, "r+") as js:
                    dados = json.load(js)
                    dados[self._estagio][self._numero] += valor
                    self._quantidade = dados[self._estagio][self._numero]
                dados = json.dumps(dados, indent=4)
            except FileNotFoundError:
                print(f"Arquivo '{self._curso}.json' não encontrado!")
            try:
                with open(caminho, "w") as js:
                    js.write(dados)
                print(f"{self._curso}: {valor} blocos do estágio " + \
                    f"{self._estagio}-{self._numero}, adicionados com sucesso!")
                print(f"Quantidade Disponível: {self._quantidade}")
                # Qualquer adição de blocos atualiza o relatorio excel
                Bloco.relatorio(True)
                self.registrarAtividade(f"Adicionou {valor} blocos de " + \
                    f"{self._curso}:{self._estagio}-{self._numero}")
                # Atividade do user é registrada
            except FileNotFoundError:
                print(f"Arquivo '{self._curso}.json' não encontrado!")
        else:
            print("Usuário precisa estar logado!")

    def retirar(self, valor):
        # Retira do atributo quantidade o valor passado como parâmetro e atualiza o json
        if self.user.log:
            # User precisa estar logado para realizar retirada
            valor = abs(int(valor))
            caminho = path.join(".", "models", f"{self.curso}.json")
            try:
                with open(caminho, "r+") as js:
                    dados = json.load(js)
                    if valor > dados[self._estagio][self._numero]:
                        print("Erro: Quantidade não disponível!")
                        return
                    dados[self._estagio][self._numero] -= valor
                    self._quantidade = dados[self._estagio][self._numero]
                    dados = json.dumps(dados, indent=4)
            except FileNotFoundError:
                print(f"Arquivo '{self._curso}.json' não encontrado!")
            try:
                with open(caminho, "w") as js:
                    js.write(dados)
                print(f"{self._curso}: {valor} blocos do estágio " + \
                    f"{self._estagio}-{self._numero}, retirados com sucesso!")
                print(f"Quantidade Disponível: {self._quantidade}")
                # Qualquer subtração de blocos atualiza o relatorio excel
                Bloco.relatorio(True)
                self.registrarAtividade(f"Retirou {valor} blocos de " + \
                    f"{self._curso}:{self._estagio}-{self._numero}")
            except FileNotFoundError:
                print(f"Arquivo '{self._curso}.json' não encontrado!")
        else:
            print("Usuário precisa estar logado!")

    def registrarAtividade(self, atividade):
        # Registra as movimentações feitas no estoque, gerando um arquivo txt por dia
        date = datetime.now().strftime("%d-%m-%Y")
        if path.isfile(path.join(".", "relatorio_de_atividades", date+".txt")):
            # Adiciona a atividade ao txt, caso arquivo já exista
            try:
                with open(path.join("relatorio_de_atividades", date+".txt"), "a") as txt:
                    time = datetime.now().strftime("%H:%M:%S")
                    txt.write(f"\n[{time}]{self.user.nome} - {atividade}")
                    # Registra horário, user e atividade
            except FileNotFoundError:
                print("Arquivo de atividades não encontrado!")
        else:
            # Caso contrário, cria um txt a adiciona a atividade
            try:
                with open(path.join(".", "relatorio_de_atividades", date+".txt"), "w") as txt:
                    time = datetime.now().strftime("%H:%M:%S")
                    txt.write(f"[{time}]{self.user.nome} - {atividade}")
                    # Registra horário, user e atividade no txt
            except FileNotFoundError:
                print("Arquivo de atividades não encontrado!")

    def __str__(self):
        return f"{self._curso} {self._estagio}-{self._numero} " + \
            f"> Quantidade Disponível: {self._quantidade}"

    def __repr__(self):
        return f"{self._curso} {self._estagio}-{self._numero} " + \
            f"> Quantidade Disponível: {self._quantidade}"


if __name__ == "__main__":
    user = User("rafael de almeida menezes", "rafaelalmeida2909@gmail.com", "123456")
    bloco = Bloco(user, "Portugues", "6a", "Livro2")
    bloco.adicionar(10)
    bloco.retirar(10)
    bloco.user.cadastrar()
    user.cadastrar()
    bloco.adicionar(10)
    bloco.retirar(10)
    bloco.user.login("rafaelalmeida2909@gmail.com", "123456")
    bloco.user.login("rafaelalmeida2909@gmail.com", "123456")
    bloco.adicionar(10)
    bloco.retirar(10)
    user1 = User("Gabriela de Almeida Menezes", "gabimenezes98@hotmail.com", "654321")
    bloco1 = Bloco(user1, "Matematica", "2a", "001")
    bloco1.user.cadastrar()
    bloco1.user.login("gabimenezes98@hotmail.com", "654321")
    bloco1.adicionar(10)
    bloco1.retirar(10)
    bloco1.user.logout()
    bloco.user.logout()
